# dataformatter.py
import openpyxl, csv

SHEET_NAME = "Object In Center"       # must be this sheet
SENSOR_COLS = [(4,5),(9,10),(14,15),(19,20)]  # (angleCol, distCol) for sensors 1..4
OFFSETS_DEG = [180.0, 180.0, 0.0, 0.0]        # sensors 1–2: +180°, 3–4: +0°
MM2IN = 1/25.4
START_ROW = 4

def convert_excel_to_sensor_format(xlsx):
    sh = openpyxl.load_workbook(xlsx, data_only=True)[SHEET_NAME]
    out, r = [], START_ROW
    while True:
        row_pts, any_data = [], False
        for s_idx,(a_col,d_col) in enumerate(SENSOR_COLS):
            a = sh.cell(row=r, column=a_col).value
            d = sh.cell(row=r, column=d_col).value
            if a is None and d is None: continue
            any_data = True
            try:
                a, d = float(a), float(d)
            except (TypeError, ValueError):
                continue
            if d == 65535: continue
            ang = ((360.0 - a) + OFFSETS_DEG[s_idx]) % 360.0
            row_pts.append([(d+10.5)*MM2IN, ang])
        if not any_data: break
        if len(row_pts) == 4: out.extend(row_pts)
        r += 1
    return out

if __name__ == "__main__":
    data = convert_excel_to_sensor_format("LiDAR_Data_Collection.xlsx")
    print(f"Processed {len(data)//4} sets")
    with open("formatted_sensor_data.csv","w",newline="") as f:
        w = csv.writer(f); w.writerow(["Test","Sensor","Distance (inches)","Angle (degrees)"])
        for i in range(0,len(data),4):
            t = i//4 + 1
            for s,(dist,ang) in enumerate(data[i:i+4], start=1):
                w.writerow([t, s, dist, ang])
